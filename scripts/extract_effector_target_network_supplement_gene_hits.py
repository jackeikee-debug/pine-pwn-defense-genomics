#!/usr/bin/env python3
"""Search route 1 candidate gene IDs and keywords in supplementary workbooks."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


FIELDS = [
    "source_id",
    "orthogroup_id",
    "matched_keyword",
    "exact_gene_match_count",
    "exact_gene_matches",
    "keyword_hit_count",
    "keyword_hit_examples",
    "searched_gene_ids",
    "workbook_path",
]


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def extract_effector_target_network_supplement_gene_hits(
    effector_target_network_path: Path,
    manifest_path: Path,
    output_path: Path,
    markdown_path: Path,
) -> list[dict[str, str]]:
    candidates = read_tsv(effector_target_network_path)
    workbook_texts = load_workbook_texts(manifest_path)
    rows = []
    for source_id, workbook_path, cells in workbook_texts:
        for candidate in candidates:
            rows.append(build_hit_row(source_id, workbook_path, cells, candidate))
    write_tsv(output_path, rows)
    write_markdown(markdown_path, rows)
    return rows


def load_workbook_texts(manifest_path: Path) -> list[tuple[str, str, list[str]]]:
    workbooks = []
    for item in read_tsv(manifest_path):
        if item.get("status") not in {"ok", "ok_from_oa_package"}:
            continue
        workbook_path = item["local_path"]
        cells = []
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells.extend(stringify(value) for value in row if stringify(value))
        workbook.close()
        workbooks.append((item["source_id"], workbook_path, cells))
    return workbooks


def build_hit_row(
    source_id: str,
    workbook_path: str,
    cells: list[str],
    candidate: dict[str, str],
) -> dict[str, str]:
    gene_ids = candidate_gene_ids(candidate)
    keyword = candidate.get("matched_keyword", "")
    exact_matches = sorted({gene_id for gene_id in gene_ids if cell_contains(cells, gene_id)})
    keyword_examples = keyword_hit_examples(cells, keyword)
    return {
        "source_id": source_id,
        "orthogroup_id": candidate["orthogroup_id"],
        "matched_keyword": keyword,
        "exact_gene_match_count": str(len(exact_matches)),
        "exact_gene_matches": ";".join(exact_matches),
        "keyword_hit_count": str(len(keyword_examples)),
        "keyword_hit_examples": " || ".join(keyword_examples[:5]),
        "searched_gene_ids": ";".join(gene_ids),
        "workbook_path": workbook_path,
    }


def candidate_gene_ids(candidate: dict[str, str]) -> list[str]:
    values = []
    for field in ["pden_gene_ids", "pmas_gene_ids", "ptab_gene_ids"]:
        values.extend(split_multi(candidate.get(field, "")))
    return sorted({strip_species_prefix(value) for value in values if value})


def strip_species_prefix(value: str) -> str:
    return value.split("|", 1)[1] if "|" in value else value


def keyword_hit_examples(cells: list[str], keyword: str) -> list[str]:
    if not keyword:
        return []
    lowered = keyword.lower()
    return [cell for cell in cells if lowered in cell.lower()]


def cell_contains(cells: list[str], needle: str) -> bool:
    needle_lower = needle.lower()
    return any(needle_lower in cell.lower() for cell in cells)


def split_multi(value: str) -> list[str]:
    return [item for item in value.replace(",", ";").split(";") if item]


def stringify(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def write_tsv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    exact_rows = [row for row in rows if row["exact_gene_match_count"] != "0"]
    keyword_rows = [row for row in rows if row["keyword_hit_count"] != "0"]
    lines = [
        "# Route 1 Supplement Gene Hits",
        "",
        "This report searches downloaded supplementary workbooks for exact route 1 candidate gene IDs and route 1 keywords.",
        "",
        f"- Candidate-source rows searched: {len(rows)}",
        f"- Rows with exact gene ID matches: {len(exact_rows)}",
        f"- Rows with keyword hits: {len(keyword_rows)}",
        "",
    ]
    for row in exact_rows[:30]:
        lines.append(
            f"- {row['source_id']} / {row['orthogroup_id']}: {row['exact_gene_matches']}"
        )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--effector_target_network", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--markdown", type=Path, required=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    rows = extract_effector_target_network_supplement_gene_hits(
        effector_target_network_path=args.effector_target_network,
        manifest_path=args.manifest,
        output_path=args.output,
        markdown_path=args.markdown,
    )
    exact = sum(row["exact_gene_match_count"] != "0" for row in rows)
    print(f"Route 1 supplement candidate-source rows searched: {len(rows)}; exact gene rows: {exact}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
