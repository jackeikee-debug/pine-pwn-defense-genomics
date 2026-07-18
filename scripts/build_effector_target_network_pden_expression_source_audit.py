#!/usr/bin/env python3
"""Audit P. densiflora public expression sources for route 1 candidate DEG support."""

from __future__ import annotations

import argparse
import csv
import io
import re
import tarfile
import zipfile
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from pypdf import PdfReader


SOURCE_FIELDS = [
    "source_id",
    "source_member",
    "material_label",
    "material_type",
    "summary_units",
    "headers",
    "detected_content",
    "transcript_level_deg_statistics",
    "effector_target_network_expression_relevance",
    "notes",
]

CANDIDATE_FIELDS = [
    "transcript_id",
    "orthogroup_ids",
    "effector_target_network_seed_symbols",
    "mapping_status",
    "best_effector_target_network_candidate_pd_gene",
    "public_transcript_deg_stats_available",
    "candidate_expression_status",
    "source_evidence_summary",
    "recommended_next_step",
    "claim_ceiling",
]

DATASET_LABELS = {
    "MOESM1": ("pden_dataset1", "Dataset 1; supplementary PDF tables and figures"),
    "MOESM2": ("pden_dataset2", "Dataset 2; Trinity transcript FASTA archive"),
    "MOESM3": ("pden_dataset3", "Dataset 3; BLAST-style Arabidopsis annotation"),
    "MOESM4": ("pden_dataset4", "Dataset 4; qRT-PCR validation R code"),
    "MOESM5": ("pden_dataset5", "Dataset 5; Trinity-to-Arabidopsis/PLAZA mapping"),
}

NO_DEG_CLAIM_CEILING = (
    "sequence-level Trinity-to-current-Pd bridge only; not exact DEG/statistical support "
    "and not direct effector-target validation"
)


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def build_effector_target_network_pden_expression_source_audit(
    oa_package: Path,
    mapping_path: Path,
    source_output_path: Path,
    candidate_output_path: Path,
    markdown_path: Path,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    source_rows = audit_oa_package(oa_package)
    write_tsv(source_output_path, source_rows, SOURCE_FIELDS)
    candidate_rows = build_candidate_expression_audit(
        mapping_path=mapping_path,
        source_rows=source_rows,
        output_path=candidate_output_path,
        markdown_path=markdown_path,
    )
    return source_rows, candidate_rows


def audit_oa_package(oa_package: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with tarfile.open(oa_package, "r:gz") as tar:
        for member in tar.getmembers():
            if "MOESM" not in member.name:
                continue
            data = tar.extractfile(member).read()
            source_id, material_label = dataset_label(member.name)
            material_type = material_type_for(member.name)
            extracted = extract_material_summary(member.name, data, material_type)
            classification = classify_source_evidence(
                material_type=material_type,
                headers=extracted["headers"].split(";") if extracted["headers"] else [],
                text=extracted["text"],
            )
            rows.append(
                {
                    "source_id": source_id,
                    "source_member": member.name,
                    "material_label": material_label,
                    "material_type": material_type,
                    "summary_units": extracted["summary_units"],
                    "headers": extracted["headers"],
                    "detected_content": classification["detected_content"],
                    "transcript_level_deg_statistics": classification["transcript_level_deg_statistics"],
                    "effector_target_network_expression_relevance": classification["effector_target_network_expression_relevance"],
                    "notes": extracted["notes"],
                }
            )
    return sorted(rows, key=lambda row: row["source_id"])


def dataset_label(member_name: str) -> tuple[str, str]:
    for token, label in DATASET_LABELS.items():
        if token in member_name:
            return label
    return "pden_unknown", "unlabeled supplementary material"


def material_type_for(member_name: str) -> str:
    suffix = Path(member_name).suffix.lower().lstrip(".")
    return suffix or "unknown"


def extract_material_summary(member_name: str, data: bytes, material_type: str) -> dict[str, str]:
    if material_type == "xlsx":
        return extract_xlsx_summary(data)
    if material_type == "docx":
        text = extract_docx_text(data)
        return {
            "summary_units": f"{len([line for line in text.splitlines() if line.strip()])} text lines",
            "headers": "",
            "text": text,
            "notes": "DOCX text extracted from word/document.xml",
        }
    if material_type == "pdf":
        text = extract_pdf_text(data)
        page_count = text.count("\f") + 1 if text else "unknown"
        return {
            "summary_units": f"{page_count} PDF pages",
            "headers": "",
            "text": text,
            "notes": "PDF text extracted with pypdf",
        }
    if material_type == "zip":
        names = extract_zip_names(data)
        return {
            "summary_units": f"{len(names)} zip members",
            "headers": "",
            "text": " ".join(names),
            "notes": ";".join(names[:10]),
        }
    return {"summary_units": "", "headers": "", "text": member_name, "notes": "unsupported material type"}


def extract_xlsx_summary(data: bytes) -> dict[str, str]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    headers: list[str] = []
    text_parts: list[str] = []
    sheet_summaries: list[str] = []
    for sheet in workbook.worksheets:
        sheet_summaries.append(f"{sheet.title}:{sheet.max_row}x{sheet.max_column}")
        values = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5), values_only=True))
        if values and not headers:
            headers = [stringify(value) for value in values[0] if stringify(value)]
        for row in values:
            text_parts.extend(stringify(value) for value in row if stringify(value))
    workbook.close()
    return {
        "summary_units": ";".join(sheet_summaries),
        "headers": ";".join(headers),
        "text": " ".join(text_parts),
        "notes": "XLSX visible header and first rows extracted with openpyxl",
    }


def extract_docx_text(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    lines = []
    for paragraph in root.findall(".//w:p", namespace):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", namespace))
        if text.strip():
            lines.append(re.sub(r"\s+", " ", text.strip()))
    return "\n".join(lines)


def extract_pdf_text(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    pages = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\f".join(pages)


def extract_zip_names(data: bytes) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        return [info.filename for info in archive.infolist()]


def classify_source_evidence(material_type: str, headers: list[str], text: str) -> dict[str, str]:
    header_text = " ".join(headers).lower()
    normalized_text = re.sub(r"\s+", " ", text).lower()
    lines = [line.lower() for line in text.splitlines() if line.strip()]
    has_transcript_deg = has_transcript_level_deg_table(header_text, lines)

    if has_transcript_deg:
        return {
            "detected_content": "transcript-level DEG statistics table",
            "transcript_level_deg_statistics": "yes",
            "effector_target_network_expression_relevance": "direct_transcript_deg_statistics_available",
        }
    if {"qseqid", "sseqid", "pident", "bitscore"}.issubset(set(split_header_tokens(headers))):
        return {
            "detected_content": "BLAST/sequence annotation table",
            "transcript_level_deg_statistics": "no",
            "effector_target_network_expression_relevance": "annotation_or_identifier_bridge_not_expression_statistics",
        }
    if "trinity id" in header_text and ("arabidopsis" in header_text or "plaza" in header_text):
        return {
            "detected_content": "Trinity-to-Arabidopsis/PLAZA identifier crosswalk",
            "transcript_level_deg_statistics": "no",
            "effector_target_network_expression_relevance": "annotation_or_identifier_bridge_not_expression_statistics",
        }
    if material_type == "zip" and "trinity.fasta" in normalized_text:
        return {
            "detected_content": "Trinity transcript FASTA archive",
            "transcript_level_deg_statistics": "no",
            "effector_target_network_expression_relevance": "sequence_resource_not_expression_statistics",
        }
    if "qrt" in normalized_text and "ngs" in normalized_text:
        return {
            "detected_content": "qRT-PCR validation code or primer context",
            "transcript_level_deg_statistics": "no",
            "effector_target_network_expression_relevance": "validation_subset_not_candidate_deg_table",
        }
    if "supplementary table s5" in normalized_text and "transcription factors" in normalized_text:
        return {
            "detected_content": "TF-family DETF summary; DEG figures and captions",
            "transcript_level_deg_statistics": "no",
            "effector_target_network_expression_relevance": "family_or_figure_level_expression_context_only",
        }
    if any(term in normalized_text for term in ["volcano", "heatmap of deg", "mapman", "deg"]):
        return {
            "detected_content": "DEG summary figure or caption context",
            "transcript_level_deg_statistics": "no",
            "effector_target_network_expression_relevance": "family_or_figure_level_expression_context_only",
        }
    return {
        "detected_content": "no effector_target_network-relevant expression statistics detected",
        "transcript_level_deg_statistics": "no",
        "effector_target_network_expression_relevance": "no_candidate_expression_statistics_detected",
    }


def has_transcript_level_deg_table(header_text: str, lines: list[str]) -> bool:
    has_id_header = any(term in header_text for term in ["trinity", "transcript", "gene_id", "gene id"])
    has_stat_header = any(term in header_text for term in ["logfc", "log2foldchange", "log2 fold", "fold change"])
    has_fdr_header = any(term in header_text for term in ["fdr", "padj", "qvalue", "adjusted p"])
    if has_id_header and has_stat_header and has_fdr_header:
        return True
    for line in lines:
        if "trinity_" not in line:
            continue
        if not any(term in line for term in ["logfc", "log2foldchange", "log2 fold", "fold change"]):
            continue
        if any(term in line for term in ["fdr", "padj", "qvalue", "adjusted p"]):
            return True
    return False


def split_header_tokens(headers: list[str]) -> list[str]:
    return [header.strip().lower() for header in headers if header.strip()]


def build_candidate_expression_audit(
    mapping_path: Path,
    source_rows: list[dict[str, str]],
    output_path: Path,
    markdown_path: Path,
) -> list[dict[str, str]]:
    public_deg_available = any(row["transcript_level_deg_statistics"] == "yes" for row in source_rows)
    source_summary = summarize_source_evidence(source_rows)
    rows = []
    for mapping in read_tsv(mapping_path):
        rows.append(candidate_expression_row(mapping, public_deg_available, source_summary))
    write_tsv(output_path, rows, CANDIDATE_FIELDS)
    write_markdown(markdown_path, rows, source_rows, source_summary, public_deg_available)
    return rows


def candidate_expression_row(
    mapping: dict[str, str],
    public_deg_available: bool,
    source_summary: str,
) -> dict[str, str]:
    mapping_status = mapping.get("mapping_status", "")
    if public_deg_available:
        expression_status = "public_deg_stats_available_candidate_lookup_required"
        next_step = "candidate_transcript_deg_lookup"
        public_deg = "yes"
    elif mapping_status == "current_pd_effector_target_network_candidate_hit":
        expression_status = "sequence_bridge_only_public_deg_stats_unavailable"
        next_step = "raw_rnaseq_reanalysis_or_author_deg_table"
        public_deg = "no"
    else:
        expression_status = "not_current_effector_target_network_candidate_public_deg_stats_unavailable"
        next_step = "deprioritize_for_current_pd_effector_target_network_exact_expression_support"
        public_deg = "no"
    return {
        "transcript_id": mapping.get("transcript_id", ""),
        "orthogroup_ids": mapping.get("orthogroup_ids", ""),
        "effector_target_network_seed_symbols": mapping.get("effector_target_network_seed_symbols", ""),
        "mapping_status": mapping_status,
        "best_effector_target_network_candidate_pd_gene": mapping.get("best_effector_target_network_candidate_pd_gene", ""),
        "public_transcript_deg_stats_available": public_deg,
        "candidate_expression_status": expression_status,
        "source_evidence_summary": source_summary,
        "recommended_next_step": next_step,
        "claim_ceiling": NO_DEG_CLAIM_CEILING,
    }


def summarize_source_evidence(source_rows: list[dict[str, str]]) -> str:
    detected = [row["detected_content"] for row in source_rows if row.get("detected_content")]
    return "; ".join(dict.fromkeys(detected))


def write_tsv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(
    path: Path,
    candidate_rows: list[dict[str, str]],
    source_rows: list[dict[str, str]],
    source_summary: str,
    public_deg_available: bool,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    status_counts = Counter(row["candidate_expression_status"] for row in candidate_rows)
    source_counts = Counter(row.get("effector_target_network_expression_relevance", "unspecified_source_context") for row in source_rows)
    lines = [
        "# Route 1 P. densiflora Expression Source Audit",
        "",
        "This report audits the available P. densiflora supplementary materials for transcript-level DEG statistics that could support route 1 Trinity-to-Pd candidate bridges.",
        "",
        f"- Supplementary materials audited: {len(source_rows)}",
        f"- Public transcript-level DEG statistics available: {'yes' if public_deg_available else 'no'}",
        f"- Candidate transcript rows summarized: {len(candidate_rows)}",
        f"- Sequence-only effector_target_network candidate bridges without public DEG statistics: {status_counts.get('sequence_bridge_only_public_deg_stats_unavailable', 0)}",
        "",
        "## Source Classes",
        "",
    ]
    for relevance, count in sorted(source_counts.items()):
        lines.append(f"- {relevance}: {count}")
    lines.extend(
        [
            "",
            "## Evidence Boundary",
            "",
            source_summary or "No effector_target_network-relevant public expression statistics were detected.",
            "",
            "The current P. densiflora bridge is therefore suitable for prioritizing raw RNA-seq reanalysis or author-table follow-up, but not for claiming exact candidate DEG support.",
        ]
    )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def stringify(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--oa-package", type=Path, required=True)
    parser.add_argument("--mapping", type=Path, required=True)
    parser.add_argument("--source-output", type=Path, required=True)
    parser.add_argument("--candidate-output", type=Path, required=True)
    parser.add_argument("--markdown", type=Path, required=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    source_rows, candidate_rows = build_effector_target_network_pden_expression_source_audit(
        oa_package=args.oa_package,
        mapping_path=args.mapping,
        source_output_path=args.source_output,
        candidate_output_path=args.candidate_output,
        markdown_path=args.markdown,
    )
    print(
        "Route 1 P. densiflora expression source audit written: "
        f"{len(source_rows)} sources; {len(candidate_rows)} candidate rows"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
