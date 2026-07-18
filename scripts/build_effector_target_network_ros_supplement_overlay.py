#!/usr/bin/env python3
"""Overlay route 1 ROS modules with downloaded expression supplement workbooks."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


FIELDS = [
    "source_id",
    "workbook_path",
    "sheet_names",
    "functional_module",
    "neighbor_symbols",
    "module_keywords",
    "support_status",
    "hit_count",
    "best_adjust_pvalue",
    "significance_status",
    "matched_terms",
    "example_rows",
    "source_url",
    "source_error",
    "claim_ceiling",
]

MODULE_KEYWORDS = {
    "superoxide_dismutase_family": [
        "superoxide dismutase",
        "superoxide",
        "SOD",
        "CSD",
        "FSD",
        "MSD",
    ],
    "ascorbate_peroxidase_ros_scavenging": [
        "ascorbate peroxidase",
        "peroxidase",
        "APX",
    ],
    "catalase_ros_scavenging": [
        "catalase",
        "CAT",
    ],
    "chloroplast_redox_gene_expression": [
        "chloroplast",
        "plastid",
        "redox",
        "oxidative stress",
        "oxidation-reduction",
        "oxidoreductase",
        "antioxidant",
        "reactive oxygen",
        "ROS",
        "PTAC",
        "FLN",
        "MRL",
    ],
    "copper_homeostasis_sod_cofactor_context": [
        "copper",
        "Cu/Zn",
        "SPL7",
    ],
    "thioredoxin_peroxiredoxin_redox": [
        "thioredoxin",
        "peroxiredoxin",
        "glutaredoxin",
        "TRX",
        "PRX",
        "CITRX",
    ],
    "unclassified_string_neighbor": [
        "oxidative stress",
        "reactive oxygen",
        "ROS",
        "oxidation-reduction",
        "oxidoreductase",
        "antioxidant",
    ],
}

CLAIM_CEILING = (
    "supplement-level module keyword or enrichment context only; not exact effector_target_network pine gene DEG "
    "or validated effector-target evidence without an identifier bridge"
)

MAX_EXAMPLES = 3
MAX_EXAMPLE_CHARS = 420
ADJUSTED_PVALUE_HEADERS = {"adjustpvalue", "adjustedpvalue", "padj", "qvalue", "fdr"}


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def build_effector_target_network_ros_supplement_overlay(
    modules_path: Path,
    manifest_path: Path,
    output_path: Path,
    markdown_path: Path,
) -> list[dict[str, str]]:
    modules = module_terms(read_tsv(modules_path))
    rows = []
    for source in read_tsv(manifest_path):
        workbook_path = Path(source.get("local_path", ""))
        if source.get("status") not in {"ok", "ok_from_oa_package"} or not workbook_path.exists():
            for module_name, terms in modules.items():
                rows.append(missing_workbook_row(source, workbook_path, module_name, terms))
            continue
        workbook_hits = scan_workbook(workbook_path, modules)
        sheet_names = workbook_sheet_names(workbook_path)
        for module_name, terms in modules.items():
            hits = workbook_hits.get(module_name, [])
            rows.append(build_overlay_row(source, workbook_path, sheet_names, module_name, terms, hits))
    write_tsv(output_path, rows)
    write_markdown(markdown_path, rows)
    return rows


def module_terms(rows: list[dict[str, str]]) -> dict[str, dict[str, list[str]]]:
    grouped: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        module_name = row.get("functional_module", "")
        partner = row.get("partner_symbol", "")
        if module_name and partner:
            grouped[module_name].add(partner)
    modules = {}
    for module_name, partners in grouped.items():
        keywords = MODULE_KEYWORDS.get(module_name, [])
        modules[module_name] = {
            "neighbor_symbols": sorted(part for part in partners if part),
            "module_keywords": keywords,
            "search_terms": sorted(set(keywords) | set(partners), key=lambda value: value.lower()),
        }
    return dict(sorted(modules.items()))


def scan_workbook(workbook_path: Path, modules: dict[str, dict[str, list[str]]]) -> dict[str, list[dict[str, str]]]:
    patterns = {
        module_name: [(term, compile_term(term)) for term in terms["search_terms"] if term]
        for module_name, terms in modules.items()
    }
    hits: dict[str, list[dict[str, str]]] = defaultdict(list)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            headers: list[str] = []
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_number == 1:
                    headers = [normalize_header(value) for value in values]
                row_text = format_row(values)
                if not row_text:
                    continue
                adjust_pvalue = extract_adjust_pvalue(headers, values)
                for module_name, module_patterns in patterns.items():
                    matched = [term for term, pattern in module_patterns if pattern.search(row_text)]
                    if matched:
                        hits[module_name].append(
                            {
                                "sheet_name": sheet.title,
                                "row_number": str(row_number),
                                "matched_terms": ";".join(sorted(set(matched), key=lambda value: value.lower())),
                                "adjust_pvalue": format_float(adjust_pvalue) if adjust_pvalue is not None else "",
                                "row_text": truncate(row_text, MAX_EXAMPLE_CHARS),
                            }
                        )
    finally:
        workbook.close()
    return hits


def workbook_sheet_names(workbook_path: Path) -> str:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return ";".join(sheet.title for sheet in workbook.worksheets)
    finally:
        workbook.close()


def build_overlay_row(
    source: dict[str, str],
    workbook_path: Path,
    sheet_names: str,
    module_name: str,
    terms: dict[str, list[str]],
    hits: list[dict[str, str]],
) -> dict[str, str]:
    matched_terms = sorted(
        {term for hit in hits for term in hit["matched_terms"].split(";") if term},
        key=lambda value: value.lower(),
    )
    best_adjust_pvalue = min_adjust_pvalue(hits)
    return {
        "source_id": source.get("source_id", ""),
        "workbook_path": str(workbook_path),
        "sheet_names": sheet_names,
        "functional_module": module_name,
        "neighbor_symbols": ";".join(terms["neighbor_symbols"]),
        "module_keywords": ";".join(terms["module_keywords"]),
        "support_status": "module_supplement_hit" if hits else "no_module_keyword_hit",
        "hit_count": str(len(hits)),
        "best_adjust_pvalue": format_float(best_adjust_pvalue) if best_adjust_pvalue is not None else "",
        "significance_status": classify_significance(hits, best_adjust_pvalue),
        "matched_terms": ";".join(matched_terms),
        "example_rows": summarize_examples(hits),
        "source_url": source.get("url", ""),
        "source_error": source.get("error", ""),
        "claim_ceiling": CLAIM_CEILING,
    }


def missing_workbook_row(
    source: dict[str, str],
    workbook_path: Path,
    module_name: str,
    terms: dict[str, list[str]],
) -> dict[str, str]:
    return {
        "source_id": source.get("source_id", ""),
        "workbook_path": str(workbook_path),
        "sheet_names": "",
        "functional_module": module_name,
        "neighbor_symbols": ";".join(terms["neighbor_symbols"]),
        "module_keywords": ";".join(terms["module_keywords"]),
        "support_status": "source_workbook_missing",
        "hit_count": "0",
        "best_adjust_pvalue": "",
        "significance_status": "",
        "matched_terms": "",
        "example_rows": "",
        "source_url": source.get("url", ""),
        "source_error": source.get("error", ""),
        "claim_ceiling": CLAIM_CEILING,
    }


def compile_term(term: str) -> re.Pattern:
    if re.search(r"\s|/|-", term):
        pattern = re.escape(term)
        flags = re.IGNORECASE
    elif term.isupper() and len(term) <= 5:
        pattern = rf"\b{re.escape(term)}[A-Z0-9_.-]*\b"
        flags = 0
    else:
        pattern = rf"\b{re.escape(term)}\b"
        flags = re.IGNORECASE
    return re.compile(pattern, flags=flags)


def format_row(values) -> str:
    return " | ".join(stringify(value) for value in values if stringify(value))


def normalize_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", stringify(value).lower())


def extract_adjust_pvalue(headers: list[str], values) -> float | None:
    for index, header in enumerate(headers):
        if header not in ADJUSTED_PVALUE_HEADERS or index >= len(values):
            continue
        value = parse_float(values[index])
        if value is not None:
            return value
    return None


def parse_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def min_adjust_pvalue(hits: list[dict[str, str]]) -> float | None:
    values = [parse_float(hit.get("adjust_pvalue", "")) for hit in hits]
    numeric = [value for value in values if value is not None]
    if not numeric:
        return None
    return min(numeric)


def classify_significance(hits: list[dict[str, str]], best_adjust_pvalue: float | None) -> str:
    if not hits:
        return ""
    if best_adjust_pvalue is None:
        return "no_adjust_pvalue_available"
    if best_adjust_pvalue < 0.05:
        return "adjusted_p_lt_0.05"
    return "adjusted_p_ge_0.05"


def format_float(value: float) -> str:
    return f"{value:.6g}"


def stringify(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def summarize_examples(hits: list[dict[str, str]]) -> str:
    examples = []
    for hit in hits[:MAX_EXAMPLES]:
        qvalue = f":q={hit['adjust_pvalue']}" if hit.get("adjust_pvalue") else ""
        examples.append(
            f"{hit['sheet_name']}:row{hit['row_number']}:{hit['matched_terms']}{qvalue}::{hit['row_text']}"
        )
    return " || ".join(examples)


def truncate(value: str, max_chars: int) -> str:
    if len(value) <= max_chars:
        return value
    return value[: max_chars - 1].rstrip() + "..."


def write_tsv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    hits = [row for row in rows if row["support_status"] == "module_supplement_hit"]
    significant_hits = [row for row in hits if row["significance_status"] == "adjusted_p_lt_0.05"]
    by_module = defaultdict(int)
    for row in hits:
        by_module[row["functional_module"]] += 1
    lines = [
        "# Route 1 ROS Supplement Overlay",
        "",
        "This report scans downloaded public expression supplementary workbooks for route 1 ROS neighbor module terms.",
        "",
        f"- Module-source rows: {len(rows)}",
        f"- Module-source workbook hits: {len(hits)}",
        f"- Module-source hits with adjusted P < 0.05: {len(significant_hits)}",
        "",
        "Evidence boundary: hits are supplement-level keyword or enrichment context, not exact route 1 pine gene DEG evidence without an identifier bridge.",
        "",
        "## Hits by Module",
        "",
    ]
    for module_name, count in sorted(by_module.items()):
        lines.append(f"- {module_name}: {count}")
    if hits:
        lines.extend(["", "## Example Hits", ""])
    for row in hits[:40]:
        lines.append(
            f"- {row['source_id']} / {row['functional_module']} "
            f"({row['matched_terms']}; {row['significance_status']}; best q={row['best_adjust_pvalue']}): "
            f"{row['example_rows']}"
        )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--modules", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--markdown", type=Path, required=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    rows = build_effector_target_network_ros_supplement_overlay(
        modules_path=args.modules,
        manifest_path=args.manifest,
        output_path=args.output,
        markdown_path=args.markdown,
    )
    hits = sum(row["support_status"] == "module_supplement_hit" for row in rows)
    print(f"Route 1 ROS supplement overlay rows written: {len(rows)}; workbook hits: {hits}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
