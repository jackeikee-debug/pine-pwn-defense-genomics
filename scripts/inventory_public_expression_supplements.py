#!/usr/bin/env python3
"""Inventory downloaded public expression supplementary workbooks."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


FIELDS = [
    "source_id",
    "workbook_path",
    "sheet_name",
    "max_row",
    "max_column",
    "headers",
    "keyword_hits",
    "example_rows",
]

DEFAULT_KEYWORDS = ["WRKY", "ERF", "MYB", "superoxide dismutase", "SOD"]


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def inventory_public_expression_supplements(
    manifest_path: Path,
    output_path: Path,
    markdown_path: Path,
    keywords: list[str] | None = None,
) -> list[dict[str, str]]:
    active_keywords = keywords or DEFAULT_KEYWORDS
    rows: list[dict[str, str]] = []
    for item in read_tsv(manifest_path):
        if item.get("status") not in {"ok", "ok_from_oa_package"}:
            continue
        workbook_path = Path(item["local_path"])
        rows.extend(inventory_workbook(item["source_id"], workbook_path, active_keywords))
    write_tsv(output_path, rows)
    write_markdown(markdown_path, rows)
    return rows


def inventory_workbook(source_id: str, workbook_path: Path, keywords: list[str]) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True))
        headers = [stringify(value) for value in values[0]] if values else []
        text = " ".join(stringify(value) for row in values for value in row)
        hits = [keyword for keyword in keywords if keyword.lower() in text.lower()]
        rows.append(
            {
                "source_id": source_id,
                "workbook_path": str(workbook_path),
                "sheet_name": sheet.title,
                "max_row": str(sheet.max_row),
                "max_column": str(sheet.max_column),
                "headers": ";".join(header for header in headers if header),
                "keyword_hits": ";".join(hits),
                "example_rows": summarize_example_rows(values[1:4]),
            }
        )
    workbook.close()
    return rows


def stringify(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def summarize_example_rows(rows) -> str:
    summaries = []
    for row in rows:
        summaries.append("|".join(stringify(value) for value in row if stringify(value)))
    return " || ".join(summary for summary in summaries if summary)


def write_tsv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    keyword_rows = [row for row in rows if row["keyword_hits"]]
    lines = [
        "# Public Expression Supplement Inventory",
        "",
        "This inventory lists downloaded workbook sheets and route 1 keyword hits in the visible sheet content.",
        "",
        f"- Workbook sheets inventoried: {len(rows)}",
        f"- Sheets with route 1 keyword hits: {len(keyword_rows)}",
        "",
    ]
    for row in keyword_rows:
        lines.append(
            f"- {row['source_id']} / {row['sheet_name']} ({row['max_row']} x {row['max_column']}): "
            f"{row['keyword_hits']}"
        )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--markdown", type=Path, required=True)
    parser.add_argument("--keywords", nargs="*", default=DEFAULT_KEYWORDS)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    rows = inventory_public_expression_supplements(
        manifest_path=args.manifest,
        output_path=args.output,
        markdown_path=args.markdown,
        keywords=args.keywords,
    )
    print(f"Supplement workbook sheets inventoried: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
