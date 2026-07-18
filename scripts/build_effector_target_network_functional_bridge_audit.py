#!/usr/bin/env python3
"""Audit indirect functional bridges between route 1 seeds and expression supplements."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


FIELDS = [
    "source_id",
    "species_id",
    "orthogroup_id",
    "effector_target_network_seed_symbols",
    "swissprot_ids",
    "tair_loci",
    "candidate_gene_ids",
    "bridge_status",
    "bridge_type",
    "matched_loci",
    "matched_expression_ids",
    "matched_expression_id_count",
    "exact_current_gene_match",
    "source_url",
    "source_error",
    "claim_ceiling",
    "recommended_next_step",
]

CLAIM_CEILING = (
    "indirect supplement-to-Arabidopsis-homolog bridge only; not a current pine gene ID match "
    "or exact candidate-gene differential-expression call"
)


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def build_effector_target_network_functional_bridge_audit(
    network_seed_path: Path,
    locus_map_path: Path,
    supplement_manifest_path: Path,
    output_path: Path,
    markdown_path: Path,
) -> list[dict[str, str]]:
    seeds = effector_target_network_seed_records(read_tsv(network_seed_path), read_tsv(locus_map_path))
    rows = []
    for source in read_tsv(supplement_manifest_path):
        workbook_path = Path(source.get("local_path", ""))
        source_matches, has_locus_columns, detected_bridge_type = scan_source_workbook(workbook_path, source)
        species_id = infer_species_id(source.get("source_id", ""))
        for seed in seeds:
            rows.append(
                build_bridge_row(
                    source,
                    species_id,
                    seed,
                    source_matches,
                    has_locus_columns,
                    detected_bridge_type,
                )
            )
    write_tsv(output_path, rows)
    write_markdown(markdown_path, rows)
    return rows


def effector_target_network_seed_records(
    seed_rows: list[dict[str, str]],
    locus_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    loci_by_swissprot = {
        row["swissprot_id"]: normalize_locus(row.get("tair_locus", ""))
        for row in locus_rows
        if row.get("swissprot_id") and row.get("tair_locus")
    }
    grouped: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for row in seed_rows:
        orthogroup_id = row.get("orthogroup_id", "")
        if not orthogroup_id:
            continue
        for field, target in [
            ("arabidopsis_homolog_symbols", "effector_target_network_seed_symbols"),
            ("arabidopsis_swissprot_ids", "swissprot_ids"),
            ("pine_gene_ids", "candidate_gene_ids"),
        ]:
            grouped[orthogroup_id][target].update(split_semicolon(row.get(field, "")))
    records = []
    for orthogroup_id, values in sorted(grouped.items()):
        swissprot_ids = sorted(values["swissprot_ids"])
        tair_loci = sorted({loci_by_swissprot[item] for item in swissprot_ids if loci_by_swissprot.get(item)})
        records.append(
            {
                "orthogroup_id": orthogroup_id,
                "effector_target_network_seed_symbols": sorted(values["effector_target_network_seed_symbols"]),
                "swissprot_ids": swissprot_ids,
                "candidate_gene_ids": sorted(values["candidate_gene_ids"]),
                "tair_loci": tair_loci,
            }
        )
    return records


def scan_source_workbook(
    workbook_path: Path,
    source: dict[str, str],
) -> tuple[dict[str, set[str]], bool, str]:
    if source.get("status") not in {"ok", "ok_from_oa_package"} or not workbook_path.exists():
        return {}, False, "none"
    matches: dict[str, set[str]] = defaultdict(set)
    has_locus_columns = False
    detected_bridge_types: set[str] = set()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            headers = []
            trinity_index = None
            arabidopsis_index = None
            sheet_bridge_type = "none"
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_number == 1:
                    headers = [normalize_header(value) for value in values]
                    trinity_index = find_header_index(headers, {"trinityid", "trinity"})
                    arabidopsis_index = find_header_index(headers, {"arabidopsisthaliana", "arabidopsis"})
                    if trinity_index is not None and arabidopsis_index is not None:
                        sheet_bridge_type = "trinity_to_arabidopsis_locus_to_effector_target_network_seed"
                    if trinity_index is None or arabidopsis_index is None:
                        trinity_index = find_header_index(headers, {"qseqid", "queryid"})
                        arabidopsis_index = find_header_index(headers, {"sseqid", "subjectid"})
                        if trinity_index is not None and arabidopsis_index is not None:
                            sheet_bridge_type = "trinity_blast_to_arabidopsis_locus_to_effector_target_network_seed"
                    has_locus_columns = has_locus_columns or (
                        trinity_index is not None and arabidopsis_index is not None
                    )
                    if sheet_bridge_type != "none":
                        detected_bridge_types.add(sheet_bridge_type)
                    continue
                if trinity_index is None or arabidopsis_index is None:
                    continue
                if max(trinity_index, arabidopsis_index) >= len(values):
                    continue
                expression_id = stringify(values[trinity_index])
                locus = normalize_locus(stringify(values[arabidopsis_index]))
                if expression_id and is_tair_locus(locus):
                    matches[locus].add(expression_id)
    finally:
        workbook.close()
    return matches, has_locus_columns, summarize_bridge_types(detected_bridge_types)


def build_bridge_row(
    source: dict[str, str],
    species_id: str,
    seed: dict[str, list[str]],
    source_matches: dict[str, set[str]],
    has_locus_columns: bool,
    detected_bridge_type: str,
) -> dict[str, str]:
    matched = {
        locus: source_matches[locus]
        for locus in seed["tair_loci"]
        if locus in source_matches
    }
    matched_expression_ids = sorted({expr for values in matched.values() for expr in values})
    return {
        "source_id": source.get("source_id", ""),
        "species_id": species_id,
        "orthogroup_id": seed["orthogroup_id"],
        "effector_target_network_seed_symbols": ";".join(seed["effector_target_network_seed_symbols"]),
        "swissprot_ids": ";".join(seed["swissprot_ids"]),
        "tair_loci": ";".join(seed["tair_loci"]),
        "candidate_gene_ids": ";".join(seed["candidate_gene_ids"]),
        "bridge_status": classify_bridge_status(source, seed, matched, has_locus_columns),
        "bridge_type": classify_bridge_type(matched, has_locus_columns, detected_bridge_type),
        "matched_loci": ";".join(sorted(matched)),
        "matched_expression_ids": ";".join(matched_expression_ids),
        "matched_expression_id_count": str(len(matched_expression_ids)),
        "exact_current_gene_match": "no",
        "source_url": source.get("url", ""),
        "source_error": source.get("error", ""),
        "claim_ceiling": CLAIM_CEILING,
        "recommended_next_step": recommended_next_step(source, species_id, matched, has_locus_columns),
    }


def classify_bridge_status(
    source: dict[str, str],
    seed: dict[str, list[str]],
    matched: dict[str, set[str]],
    has_locus_columns: bool,
) -> str:
    if source.get("status") not in {"ok", "ok_from_oa_package"}:
        return "source_workbook_missing"
    if not seed["tair_loci"]:
        return "blocked_missing_arabidopsis_locus_map"
    if matched:
        return "indirect_arabidopsis_locus_bridge_found"
    if has_locus_columns:
        return "no_effector_target_network_locus_match"
    return "source_has_no_arabidopsis_locus_columns"


def classify_bridge_type(matched: dict[str, set[str]], has_locus_columns: bool, detected_bridge_type: str) -> str:
    if matched:
        return detected_bridge_type if detected_bridge_type != "none" else "arabidopsis_locus_to_effector_target_network_seed"
    if has_locus_columns:
        return "arabidopsis_locus_scan_no_match"
    return "none"


def summarize_bridge_types(values: set[str]) -> str:
    if not values:
        return "none"
    if len(values) == 1:
        return next(iter(values))
    return "mixed:" + ";".join(sorted(values))


def recommended_next_step(
    source: dict[str, str],
    species_id: str,
    matched: dict[str, set[str]],
    has_locus_columns: bool,
) -> str:
    if matched:
        if species_id == "pden":
            return (
                "Recover the corresponding DEG/statistics table for these Trinity IDs, then align or map "
                "Trinity transcripts back to current Pd proteins before claiming exact candidate expression."
            )
        return "Treat as an indirect functional bridge and seek a current species-specific protein/transcript ID map."
    if source.get("status") not in {"ok", "ok_from_oa_package"}:
        return "Recover the missing supplement workbook before bridge testing."
    if not has_locus_columns and species_id == "pmas":
        return "Locate PITA-to-gmmutg/STRG mapping or align PITA transcript/protein sequences to the current pmas proteome."
    return "No route 1 locus bridge found in this source; keep as background unless a compatible ID map is recovered."


def split_semicolon(value: str) -> set[str]:
    return {item.strip() for item in value.split(";") if item.strip()}


def normalize_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", stringify(value).lower())


def find_header_index(headers: list[str], candidates: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in candidates:
            return index
    return None


def normalize_locus(value: str) -> str:
    return stringify(value).upper().rstrip(";")


def is_tair_locus(value: str) -> bool:
    return bool(re.fullmatch(r"AT[1-5CM]G\d{5}", value))


def stringify(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def infer_species_id(source_id: str) -> str:
    if source_id.startswith("pden"):
        return "pden"
    if source_id.startswith("pmas"):
        return "pmas"
    if source_id.startswith("ptab"):
        return "ptab"
    return "unknown"


def write_tsv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    bridge_rows = [row for row in rows if row["bridge_status"] == "indirect_arabidopsis_locus_bridge_found"]
    lines = [
        "# Route 1 Functional Bridge Audit",
        "",
        "This audit checks whether downloaded expression supplements can be indirectly bridged to route 1 seeds through Arabidopsis homolog loci.",
        "",
        f"- Source-orthogroup rows: {len(rows)}",
        f"- Indirect Arabidopsis locus bridges found: {len(bridge_rows)}",
        "",
        "Evidence boundary: bridges are not current pine gene-ID matches and do not by themselves prove exact candidate-gene differential expression.",
        "",
    ]
    if bridge_rows:
        lines.extend(["## Indirect Bridges", ""])
    for row in bridge_rows:
        lines.append(
            f"- {row['source_id']} / {row['orthogroup_id']} ({row['effector_target_network_seed_symbols']}): "
            f"{row['matched_loci']} -> {row['matched_expression_id_count']} transcript IDs"
        )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--network-seed", type=Path, required=True)
    parser.add_argument("--locus-map", type=Path, required=True)
    parser.add_argument("--supplement-manifest", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--markdown", type=Path, required=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    rows = build_effector_target_network_functional_bridge_audit(
        network_seed_path=args.network_seed,
        locus_map_path=args.locus_map,
        supplement_manifest_path=args.supplement_manifest,
        output_path=args.output,
        markdown_path=args.markdown,
    )
    bridges = sum(row["bridge_status"] == "indirect_arabidopsis_locus_bridge_found" for row in rows)
    print(f"Route 1 functional bridge audit rows written: {len(rows)}; indirect bridges: {bridges}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
